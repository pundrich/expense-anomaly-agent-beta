"""
Excel report writer.

Builds three sheets:
  Summary             - run metadata + flag color counts (formulas, not hardcoded)
  Flagged Transactions - per-flag detail with explanation + classification,
                         row-tinted by flag color
  Category Stats      - per-category count / mean / std used by the detector

Color convention follows the financial-modeling guide:
  - blue  text  for hardcoded inputs (run metadata)
  - black text  for formulas
  - flag tints applied as fill on the Flagged sheet only
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROFESSIONAL_FONT = "Arial"

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=PROFESSIONAL_FONT, bold=True, color="FFFFFF", size=11)

INPUT_FONT = Font(name=PROFESSIONAL_FONT, color="0000FF")
FORMULA_FONT = Font(name=PROFESSIONAL_FONT, color="000000")
BODY_FONT = Font(name=PROFESSIONAL_FONT, color="000000")
TITLE_FONT = Font(name=PROFESSIONAL_FONT, bold=True, size=14, color="1F4E78")

FLAG_FILL = {
    "RED":    PatternFill("solid", start_color="F8CBAD"),
    "YELLOW": PatternFill("solid", start_color="FFE699"),
    "GREEN":  PatternFill("solid", start_color="C6E0B4"),
}
FLAG_FONT = {
    "RED":    Font(name=PROFESSIONAL_FONT, bold=True, color="9C0006"),
    "YELLOW": Font(name=PROFESSIONAL_FONT, bold=True, color="9C5700"),
    "GREEN":  Font(name=PROFESSIONAL_FONT, bold=True, color="006100"),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '"$"#,##0.00;[Red]("$"#,##0.00);-'
PERCENT_FMT = '0.0%;[Red]-0.0%;-'
DEVIATION_FMT = '+0%;[Red]-0%;-'


def _apply_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _autosize(ws, min_w: int = 10, max_w: int = 48) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        widest = max((len(str(c.value)) if c.value is not None else 0)
                     for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(max_w, widest + 2))


def write_report(
    flagged: pd.DataFrame,
    explanations: dict[str, str],
    classifications: dict[str, dict],
    stats: pd.DataFrame,
    total_txns: int,
    threshold: float,
    out_path: Path,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    _write_flagged_sheet(wb, flagged, explanations, classifications)
    _write_summary_sheet(wb, total_txns, len(flagged), threshold)
    _write_stats_sheet(wb, stats)

    # reorder so Summary is first
    wb.move_sheet("Summary", offset=-2)

    wb.save(out_path)


# --- sheets --------------------------------------------------------------

def _write_summary_sheet(wb: Workbook, total: int, n_flagged: int, threshold: float) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Expense Anomaly Agent - Run Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    rows = [
        ("Total transactions reviewed", total, "input"),
        ("Z-score threshold (σ)", threshold, "input"),
        ("Flagged transactions", n_flagged, "input"),
        ("Flag rate", "=B4/B2", "formula"),
        ("RED flags",    "=COUNTIF('Flagged Transactions'!N:N,\"RED\")", "formula"),
        ("YELLOW flags", "=COUNTIF('Flagged Transactions'!N:N,\"YELLOW\")", "formula"),
        ("GREEN flags",  "=COUNTIF('Flagged Transactions'!N:N,\"GREEN\")", "formula"),
    ]
    for i, (label, value, kind) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(name=PROFESSIONAL_FONT, bold=True)
        c = ws.cell(row=i, column=2, value=value)
        c.font = INPUT_FONT if kind == "input" else FORMULA_FONT
        if label == "Flag rate":
            c.number_format = PERCENT_FMT

    # tidy widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16


def _write_flagged_sheet(
    wb: Workbook,
    flagged: pd.DataFrame,
    explanations: dict[str, str],
    classifications: dict[str, dict],
) -> None:
    ws = wb.active
    ws.title = "Flagged Transactions"

    headers = [
        "Transaction ID", "Date", "Requester", "Department", "Vendor",
        "Category", "Amount", "Cat. Mean", "Expected Max",
        "Deviation %", "Z-score", "Explanation", "Method", "Flag", "Rationale",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        _apply_header(c)
    ws.freeze_panes = "A2"

    for i, (_, r) in enumerate(flagged.iterrows(), start=2):
        txn = r["transaction_id"]
        cls = classifications.get(txn, {"flag": "YELLOW", "rationale": "", "method": ""})
        explanation = explanations.get(txn, "")

        ws.cell(row=i, column=1, value=txn).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["date"]).font = BODY_FONT
        ws.cell(row=i, column=3, value=r["requester"]).font = BODY_FONT
        ws.cell(row=i, column=4, value=r["department"]).font = BODY_FONT
        ws.cell(row=i, column=5, value=r["vendor"]).font = BODY_FONT
        ws.cell(row=i, column=6, value=r["category"]).font = BODY_FONT

        amt = ws.cell(row=i, column=7, value=float(r["amount"]))
        amt.number_format = CURRENCY_FMT
        amt.font = INPUT_FONT  # source data

        mean = ws.cell(row=i, column=8, value=float(r["cat_mean"]))
        mean.number_format = CURRENCY_FMT
        mean.font = INPUT_FONT

        emax = ws.cell(row=i, column=9, value=float(r["expected_max"]))
        emax.number_format = CURRENCY_FMT
        emax.font = INPUT_FONT

        # deviation as a formula so the cell stays live if amounts are edited
        dev = ws.cell(row=i, column=10, value=f"=(G{i}-H{i})/H{i}")
        dev.number_format = DEVIATION_FMT
        dev.font = FORMULA_FONT

        z = ws.cell(row=i, column=11, value=float(r["z_score"]))
        z.number_format = "0.00"
        z.font = INPUT_FONT

        exp = ws.cell(row=i, column=12, value=explanation)
        exp.font = BODY_FONT
        exp.alignment = Alignment(wrap_text=True, vertical="top")

        method = ws.cell(row=i, column=13, value=cls.get("method", ""))
        method.font = BODY_FONT

        flag = cls.get("flag", "YELLOW").upper()
        flag_cell = ws.cell(row=i, column=14, value=flag)
        flag_cell.font = FLAG_FONT.get(flag, BODY_FONT)
        flag_cell.alignment = Alignment(horizontal="center")

        rat = ws.cell(row=i, column=15, value=cls.get("rationale", ""))
        rat.font = BODY_FONT
        rat.alignment = Alignment(wrap_text=True, vertical="top")

        # tint the row by flag color (light shade)
        if flag in FLAG_FILL:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = FLAG_FILL[flag]
                ws.cell(row=i, column=col).border = BORDER

    _autosize(ws, min_w=12, max_w=42)
    # explanation + rationale want to be wider
    ws.column_dimensions["L"].width = 48
    ws.column_dimensions["O"].width = 48
    # row height for wrapped text
    for i in range(2, len(flagged) + 2):
        ws.row_dimensions[i].height = 42


def _write_stats_sheet(wb: Workbook, stats: pd.DataFrame) -> None:
    ws = wb.create_sheet("Category Stats")
    headers = ["Category", "Count", "Mean", "Std Dev", "Median", "Expected Max (μ+2σ)"]
    for col, h in enumerate(headers, start=1):
        _apply_header(ws.cell(row=1, column=col, value=h))

    for i, (_, r) in enumerate(stats.iterrows(), start=2):
        ws.cell(row=i, column=1, value=r["category"]).font = BODY_FONT
        ws.cell(row=i, column=2, value=int(r["cat_n"])).font = INPUT_FONT
        for col_idx, key in enumerate(("cat_mean", "cat_std", "cat_median"), start=3):
            c = ws.cell(row=i, column=col_idx, value=float(r[key]))
            c.number_format = CURRENCY_FMT
            c.font = INPUT_FONT
        emax = ws.cell(row=i, column=6, value=f"=C{i}+2*D{i}")
        emax.number_format = CURRENCY_FMT
        emax.font = FORMULA_FONT

    _autosize(ws, min_w=14, max_w=28)
